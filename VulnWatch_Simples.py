import requests
import datetime
import urllib.parse
import json
import re
import pandas as pd
import openpyxl

Vendor_list = []

def Get_vendors():
    # Load NVD data from a local .json file
    dataframe  = openpyxl.load_workbook('vendors_list.xlsx')
    print("dd")

    # Define variable to read sheet
    dataframe1 = dataframe.active

    for row  in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            print(col[row].value)
            Vendor_list.append(col[row].value)
    return Vendor_list
Get_vendors()
print(Vendor_list)