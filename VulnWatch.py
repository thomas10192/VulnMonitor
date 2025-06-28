import requests
import datetime
import json
import re
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
from dotenv import load_dotenv 

load_dotenv() 

def send_email(subject, body, to_emails, from_email, smtp_server, smtp_port, username, password):
    msg = MIMEMultipart()
    msg["From"] = from_email
    msg["To"] = ", ".join(to_emails)  # Displayed in email client
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(username, password)
        server.send_message(msg)
        server.quit()
        print("Email sent successfully!")
    except Exception as e:
        print(f"Failed to send email: {e}")


Vendor_list = []

def Get_vendors():
    # Load NVD data from a local .json file
    dataframe  = openpyxl.load_workbook('vendors_list.xlsx')
    
    # Define variable to read sheet
    dataframe1 = dataframe.active

    for row  in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            Vendor_list.append(col[row].value)
    return Vendor_list


def get_cvss_metrics(metrics):
    cvss_data = None
    vector_version = None

    if "cvssMetricV31" in metrics and metrics["cvssMetricV31"]:
        cvss_data = metrics["cvssMetricV31"][0]["cvssData"]

    elif "cvssMetricV40" in metrics and metrics["cvssMetricV40"]:
        cvss_data = metrics["cvssMetricV40"][0]["cvssData"]
    
    return cvss_data

def check_descriptions_language(description_data):
    description_info = None
    no_desc = "No description"
    
    for desc in description_data:
        
        if desc.get("lang") == "en":
            return desc.get("value", "")
    return no_desc


vendors_list = [v.strip() for v in Get_vendors() if v]

def cve_mentions_vendor(cve_entry):
    texts = []

    # 1. Descriptions
    for desc in cve_entry.get("descriptions", []):
        texts.append(desc.get("value", ""))

    # 2. References (URLs or sources might contain vendor names)
    for ref in cve_entry.get("references", []):
        texts.append(ref.get("url", ""))
        texts.append(ref.get("source", ""))

    # 3. (Optional) Metrics source
    metrics = cve_entry.get("metrics", {})
    for key in metrics:
        for metric_item in metrics[key]:
            texts.append(metric_item.get("source", ""))

    #print(texts)
    # Combine all text and search
    combined_text = " ".join(texts)
    combined_text_lower = combined_text.lower()
   
    mentioned_vendors = [vendor for vendor in vendors_list if vendor.lower() in combined_text_lower]
    
    return bool(mentioned_vendors)


def fetch_recent_critical_cves_from_history():
   # 1. Calculate UTC timestamps for now and 24 hours ago
    now = datetime.datetime.utcnow()
    start_time = now - datetime.timedelta(days=1)

    # Format to required ISO 8601 format: YYYY-MM-DDTHH:MM:SSZ
    def format_iso8601_z(dt):
        return dt.strftime("%Y-%m-%dT%H:%M:%SZ")

    start_str = format_iso8601_z(start_time)
    end_str = format_iso8601_z(now)

    # 2. Fetch CVE history for that range
    history_url = "https://services.nvd.nist.gov/rest/json/cves/2.0"
    params = {
        "pubStartDate": start_str,
        "pubEndDate": end_str
    }

    response = requests.get(history_url, params=params)
    

    if response.status_code != 200:
        print(f"Failed to retrive data {response.status_code}")
        exit()
        

    data = response.json()

    

    # Serializing json
    json_object = json.dumps(data, indent=4)
    

    # Writing to sample.json
    with open("nvdcve-1.1-recent.json", "w") as outfile:
       outfile.write(json_object)

    # Load NVD data from a local .json file
    with open("nvdcve-1.1-recent.json", "r") as file:
        data = json.load(file)

    
    #print(f"🔍 Checking CVEs updated between {start_time} and {now}")
    print("  ")

    return data

def main():
    email_body = []
    cve_count = 0
    data = fetch_recent_critical_cves_from_history()

    for item in data.get("vulnerabilities", []):
        cve_data = item["cve"]
        cve_id = cve_data["id"]
        metrics = cve_data["metrics"]
        descriptions = cve_data["descriptions"]
        GetcvssMetric = get_cvss_metrics(metrics)

        if GetcvssMetric is None:
            continue

        baseSeverity = GetcvssMetric["baseSeverity"]
        if baseSeverity == "CRITICAL" and cve_mentions_vendor(cve_data):
            
            cve_count += 1
            details = f"""
Relevant CVE: {cve_id}
Published Date: {cve_data["published"]}
Modified Date: {cve_data["lastModified"]}
Severity: {baseSeverity}
Base Score: {GetcvssMetric["baseScore"]}
Vector: {GetcvssMetric["vectorString"]}
Description: {check_descriptions_language(descriptions)}

"""
            print(details)
            email_body.append(details)

    if cve_count == 0:
        message = "No new CVEs for today!!!"
        print(message)
        email_body.append(message)
    else:
        header = f"There are {cve_count} CVEs to look at!\n\n"
        print(header)
        email_body.insert(0, header)

    # Send email inside main
    send_email(
        subject="Daily CVE Report",
        body="".join(email_body),
        to_emails = os.getenv("EMAIL_TO").split(","),
        from_email = os.getenv("EMAIL_USER"),
        smtp_server = "smtp.gmail.com",
        smtp_port = 587,
        username = os.getenv("EMAIL_USER"),
        password = os.getenv("EMAIL_PASS")
    )


if __name__ == "__main__":
    main()


